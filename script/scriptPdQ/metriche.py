import requests
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
from dotenv import load_dotenv
import os

load_dotenv()

GITHUB_TOKEN = ""
ORGANIZATION = "GlitchHub-Team"
PROJECT_NUMBER = 2
OUTPUT_FILE = "report_sprint_gerarchico.xlsx"

PR_REPOS = [
    "GlitchHub-Team.github.io",
    
]

EXPECTED_HOURS_FIELD = "Expected Worked Hours"
WORKED_HOURS_FIELD = "Worked hours"
ROLE_FIELD_NAME = "Sprint role"

EXCLUDED_LABELS = ["task-palestra"]

STANDARD_ROLES = [
    "Responsabile",
    "Amministratore",
    "Analista",
    "Progettista",
    "Programmatore",
    "Verificatore"
]

HOURLY_RATES = {
    "Responsabile": 30,
    "Amministratore": 20,
    "Analista": 25,
    "Progettista": 25,
    "Programmatore": 15,
    "Verificatore": 15,
    "Non assegnato": 15
}

BAT_BUDGET = 12975


def get_all_project_items(org, project_num, token):
    url = "https://api.github.com/graphql"
    headers = {"Authorization": f"bearer {token}"}

    all_nodes = []
    has_next_page = True
    end_cursor = None

    while has_next_page:
        after_arg = f', after: "{end_cursor}"' if end_cursor else ""

        query = """
        query {
          organization(login: "%s") {
            projectV2(number: %d) {
              items(first: 100%s) {
                pageInfo { hasNextPage, endCursor }
                nodes {
                  content {
                    ... on Issue {
                      __typename
                      title
                      number
                      state
                      url
                      createdAt
                      author { login }
                      assignees(first: 10) {
                        nodes { login }
                      }
                      labels(first: 10) { nodes { name } }
                    }
                    ... on PullRequest {
                      __typename
                      title
                      number
                      state
                      url
                      createdAt
                      mergedAt
                      closedAt
                      author { login }
                      assignees(first: 10) {
                        nodes { login }
                      }
                      labels(first: 10) { nodes { name } }
                    }
                  }
                  fieldValues(first: 20) {
                    nodes {
                      ... on ProjectV2ItemFieldIterationValue {
                        title
                        startDate
                        duration
                        field { ... on ProjectV2IterationField { name } }
                      }
                      ... on ProjectV2ItemFieldSingleSelectValue {
                        name
                        field { ... on ProjectV2SingleSelectField { name } }
                      }
                      ... on ProjectV2ItemFieldNumberValue {
                        number
                        field { ... on ProjectV2Field { name } }
                      }
                      ... on ProjectV2ItemFieldTextValue {
                        text
                        field { ... on ProjectV2Field { name } }
                      }
                    }
                  }
                }
              }
            }
          }
        }
        """ % (org, project_num, after_arg)

        response = requests.post(url, json={'query': query}, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Errore HTTP: {response.status_code} - {response.text}")

        data = response.json()
        if 'errors' in data:
            raise Exception(data['errors'][0]['message'])

        chunk = data['data']['organization']['projectV2']['items']
        all_nodes.extend(chunk['nodes'])
        has_next_page = chunk['pageInfo']['hasNextPage']
        end_cursor = chunk['pageInfo']['endCursor']
        time.sleep(0.2)

    return all_nodes


def get_all_prs_from_repos(org, repos, token):
    """Recupera TUTTE le Pull Request dai repository specificati (aperte, chiuse, merged)."""
    url = "https://api.github.com/graphql"
    headers = {"Authorization": f"bearer {token}"}
    all_prs = []

    for repo in repos:
        has_next_page = True
        end_cursor = None

        while has_next_page:
            after_arg = f', after: "{end_cursor}"' if end_cursor else ""

            query = """
            query {
              repository(owner: "%s", name: "%s") {
                pullRequests(first: 100, states: [OPEN, CLOSED, MERGED], orderBy: {field: CREATED_AT, direction: DESC}%s) {
                  pageInfo { hasNextPage, endCursor }
                  nodes {
                    title
                    number
                    state
                    url
                    createdAt
                    mergedAt
                    closedAt
                    author { login }
                    assignees(first: 10) {
                      nodes { login }
                    }
                    labels(first: 10) { nodes { name } }
                    reviews(first: 1) {
                      totalCount
                    }
                  }
                }
              }
            }
            """ % (org, repo, after_arg)

            response = requests.post(url, json={'query': query}, headers=headers)
            if response.status_code != 200:
                print(f"Errore HTTP per repo {repo}: {response.status_code}")
                break

            data = response.json()
            if 'errors' in data:
                print(f"Errore GraphQL per repo {repo}: {data['errors'][0]['message']}")
                break

            pr_chunk = data['data']['repository']['pullRequests']
            for pr in pr_chunk['nodes']:
                pr['_repo'] = repo  
            all_prs.extend(pr_chunk['nodes'])
            has_next_page = pr_chunk['pageInfo']['hasNextPage']
            end_cursor = pr_chunk['pageInfo']['endCursor']
            time.sleep(0.2)

    return all_prs


def format_datetime(iso_str):
    """Formatta una data ISO in formato leggibile."""
    if not iso_str:
        return ""
    from datetime import datetime
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return iso_str


def calc_duration_days(start_iso, end_iso):
    """Calcola la durata in giorni tra due date ISO."""
    if not start_iso or not end_iso:
        return None
    from datetime import datetime
    try:
        start = datetime.fromisoformat(start_iso.replace("Z", "+00:00"))
        end = datetime.fromisoformat(end_iso.replace("Z", "+00:00"))
        delta = end - start
        return round(delta.total_seconds() / 86400, 1)
    except Exception:
        return None


def get_all_repo_prs(org, repos, token):
    """Recupera TUTTE le Pull Request dai repository specificati."""
    url = "https://api.github.com/graphql"
    headers = {"Authorization": f"bearer {token}"}
    all_prs = []

    for repo_name in repos:
        has_next_page = True
        end_cursor = None

        while has_next_page:
            after_arg = f', after: "{end_cursor}"' if end_cursor else ""

            query = """
            query {
              repository(owner: "%s", name: "%s") {
                pullRequests(first: 100, orderBy: {field: CREATED_AT, direction: DESC}%s) {
                  pageInfo { hasNextPage, endCursor }
                  nodes {
                    title
                    number
                    state
                    url
                    createdAt
                    mergedAt
                    closedAt
                    author { login }
                    assignees(first: 10) {
                      nodes { login }
                    }
                    labels(first: 10) { nodes { name } }
                    reviews(first: 1) {
                      totalCount
                    }
                  }
                }
              }
            }
            """ % (org, repo_name, after_arg)

            response = requests.post(url, json={'query': query}, headers=headers)
            if response.status_code != 200:
                print(f"Errore HTTP per repo {repo_name}: {response.status_code}")
                break

            data = response.json()
            if 'errors' in data:
                print(f"Errore GraphQL per repo {repo_name}: {data['errors'][0]['message']}")
                break

            pr_chunk = data['data']['repository']['pullRequests']
            for pr in pr_chunk['nodes']:
                pr['_repo'] = repo_name
            all_prs.extend(pr_chunk['nodes'])
            has_next_page = pr_chunk['pageInfo']['hasNextPage']
            end_cursor = pr_chunk['pageInfo']['endCursor']
            time.sleep(0.2)

    return all_prs


def main():
    GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")
    try:
        items = get_all_project_items(ORGANIZATION, PROJECT_NUMBER, GITHUB_TOKEN)
    except Exception:
        return

    sprint_summary = {}
    role_hours_expected = defaultdict(float)
    role_hours_worked = defaultdict(float)
    sprint_role_hours = defaultdict(lambda: defaultdict(lambda: {'expected': 0, 'worked': 0}))
    author_hours_expected = defaultdict(float)
    author_hours_worked = defaultdict(float)
    sprint_author_hours = defaultdict(lambda: defaultdict(lambda: {'expected': 0, 'worked': 0}))

    assignee_hours = defaultdict(lambda: {'expected': 0, 'worked': 0, 'issue_count': 0})
    sprint_assignee_hours = defaultdict(lambda: defaultdict(lambda: {'expected': 0, 'worked': 0, 'issue_count': 0}))

    
    sprint_dates = {}

    
    sprint_palestra_hours = defaultdict(float)

    
    pr_data = []
    if PR_REPOS:
        try:
            all_repo_prs = get_all_prs_from_repos(ORGANIZATION, PR_REPOS, GITHUB_TOKEN)
        except Exception as e:
            print(f"Errore recupero PR dai repo: {e}")
            all_repo_prs = []

        for pr in all_repo_prs:
            labels_pr = [l['name'].lower() for l in pr.get('labels', {}).get('nodes', [])]
            if any(excl.lower() in labels_pr for excl in EXCLUDED_LABELS):
                continue

            pr_end = pr.get('mergedAt') or pr.get('closedAt')
            duration = calc_duration_days(pr.get('createdAt'), pr_end)

            pr_author = (pr.get('author') or {}).get('login', 'Sconosciuto')
            pr_assignees = [a.get('login', 'Sconosciuto') for a in pr.get('assignees', {}).get('nodes', [])]

            pr_data.append({
                'number': pr['number'],
                'title': pr['title'],
                'author': pr_author,
                'assignees': pr_assignees,
                'state': pr['state'],
                'created_at': pr.get('createdAt'),
                'merged_at': pr.get('mergedAt'),
                'closed_at': pr.get('closedAt'),
                'end_date': pr_end,
                'duration_days': duration,
                'repo': pr.get('_repo', ''),
                'url': pr.get('url', ''),
            })

    for item in items:
        if not item['content']:
            continue

        content = item['content']
        content_type = content.get('__typename', 'Issue')
        is_pr = content_type == 'PullRequest'

        labels = [l['name'].lower() for l in content['labels']['nodes']]

        is_palestra = any(excl.lower() in labels for excl in EXCLUDED_LABELS)

        
        if is_pr:
            continue

        sprint_name = "Backlog"
        expected_hours = 0
        worked_hours = 0
        role = "Non assegnato"

        for fv in item['fieldValues']['nodes']:
            if not fv or 'field' not in fv:
                continue

            field_name = fv['field'].get('name', '')

            if field_name == 'Sprint':
                sprint_name = fv.get('title') or fv.get('name') or "Backlog"
                
                start_date = fv.get('startDate')
                duration_days = fv.get('duration')
                if start_date and duration_days and sprint_name not in sprint_dates:
                    from datetime import datetime, timedelta
                    try:
                        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                        end_dt = start_dt + timedelta(days=duration_days)
                        sprint_dates[sprint_name] = {
                            'start': start_dt,
                            'end': end_dt
                        }
                    except Exception:
                        pass

            if field_name == EXPECTED_HOURS_FIELD:
                expected_hours = fv.get('number') or 0

            if field_name == WORKED_HOURS_FIELD:
                worked_hours = fv.get('number') or 0

            if field_name == ROLE_FIELD_NAME:
                role = fv.get('name') or fv.get('text') or "Non assegnato"

        
        if is_palestra:
            sprint_palestra_hours[sprint_name] += worked_hours
            continue

        author = content.get('author', {})
        author_login = author.get('login', 'Sconosciuto') if author else 'Sconosciuto'

        state = content['state']

        if sprint_name not in sprint_summary:
            sprint_summary[sprint_name] = {
                'completate': 0,
                'ritardo': 0,
                'ore_stimate': 0,
                'ore_lavorate': 0,
                'ore_stimate_completate': 0,
                'ore_lavorate_completate': 0
            }

        sprint_summary[sprint_name]['ore_stimate'] += expected_hours
        sprint_summary[sprint_name]['ore_lavorate'] += worked_hours

        if state in ('CLOSED', 'MERGED'):
            sprint_summary[sprint_name]['completate'] += 1
            sprint_summary[sprint_name]['ore_stimate_completate'] += expected_hours
            sprint_summary[sprint_name]['ore_lavorate_completate'] += worked_hours

        if state in ('CLOSED', 'MERGED') and 'delayed' in labels:
            sprint_summary[sprint_name]['ritardo'] += 1

        role_hours_expected[role] += expected_hours
        role_hours_worked[role] += worked_hours
        sprint_role_hours[sprint_name][role]['expected'] += expected_hours
        sprint_role_hours[sprint_name][role]['worked'] += worked_hours

        author_hours_expected[author_login] += expected_hours
        author_hours_worked[author_login] += worked_hours
        sprint_author_hours[sprint_name][author_login]['expected'] += expected_hours
        sprint_author_hours[sprint_name][author_login]['worked'] += worked_hours

        assignees = content.get('assignees', {}).get('nodes', [])
        if assignees:
            num_assignees = len(assignees)
            hours_per_assignee_exp = expected_hours / num_assignees if num_assignees > 0 else 0
            hours_per_assignee_wrk = worked_hours / num_assignees if num_assignees > 0 else 0

            for assignee in assignees:
                assignee_login = assignee.get('login', 'Sconosciuto')
                assignee_hours[assignee_login]['expected'] += hours_per_assignee_exp
                assignee_hours[assignee_login]['worked'] += hours_per_assignee_wrk
                assignee_hours[assignee_login]['issue_count'] += 1
                sprint_assignee_hours[sprint_name][assignee_login]['expected'] += hours_per_assignee_exp
                sprint_assignee_hours[sprint_name][assignee_login]['worked'] += hours_per_assignee_wrk
                sprint_assignee_hours[sprint_name][assignee_login]['issue_count'] += 1
        else:
            assignee_hours['Non assegnato']['expected'] += expected_hours
            assignee_hours['Non assegnato']['worked'] += worked_hours
            assignee_hours['Non assegnato']['issue_count'] += 1
            sprint_assignee_hours[sprint_name]['Non assegnato']['expected'] += expected_hours
            sprint_assignee_hours[sprint_name]['Non assegnato']['worked'] += worked_hours
            sprint_assignee_hours[sprint_name]['Non assegnato']['issue_count'] += 1

    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    role_font = Font(bold=True)

    ws = wb.active
    ws.title = 'Report'

    def sprint_sort_key(name):
        import re
        match = re.search(r'(\d+)', name)
        return int(match.group(1)) if match else 999

    sorted_sprints = sorted(
        [s for s in sprint_role_hours.keys() if s != "Backlog"],
        key=sprint_sort_key
    )

    current_row = 1

    title_font = Font(bold=True, size=14)
    ws.cell(row=current_row, column=1, value="RIEPILOGO ORE PER RUOLO").font = title_font
    current_row += 2

    header_row = ['Ruolo']
    for sprint in sorted_sprints:
        sprint_short = sprint.replace("Sprint ", "S")
        header_row.extend([f'{sprint_short} Prev.', f'{sprint_short} Eff.'])

    for col, val in enumerate(header_row, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row += 1

    role_totals = {}
    for role in STANDARD_ROLES:
        ws.cell(row=current_row, column=1, value=role).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border
        col = 2
        role_total_exp = 0
        role_total_wrk = 0
        for sprint in sorted_sprints:
            exp = sprint_role_hours[sprint][role]['expected']
            wrk = sprint_role_hours[sprint][role]['worked']
            role_total_exp += exp
            role_total_wrk += wrk
            ws.cell(row=current_row, column=col, value=exp).border = thin_border
            ws.cell(row=current_row, column=col+1, value=wrk).border = thin_border
            col += 2
        role_totals[role] = {'expected': role_total_exp, 'worked': role_total_wrk}
        current_row += 1

    total_fill = PatternFill('solid', start_color='D9E2F3')
    ws.cell(row=current_row, column=1, value="TOTALE").font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = total_fill
    ws.cell(row=current_row, column=1).border = thin_border
    col = 2
    grand_total_exp = 0
    grand_total_wrk = 0
    for sprint in sorted_sprints:
        total_exp = sum(sprint_role_hours[sprint][r]['expected'] for r in STANDARD_ROLES)
        total_wrk = sum(sprint_role_hours[sprint][r]['worked'] for r in STANDARD_ROLES)
        grand_total_exp += total_exp
        grand_total_wrk += total_wrk
        cell_exp = ws.cell(row=current_row, column=col, value=total_exp)
        cell_wrk = ws.cell(row=current_row, column=col+1, value=total_wrk)
        for c in [cell_exp, cell_wrk]:
            c.font = Font(bold=True)
            c.fill = total_fill
            c.border = thin_border
        col += 2

    tot_row = 11
    ws.cell(row=tot_row, column=1, value="TOTALI").font = title_font
    tot_row += 1

    for col, val in enumerate(['Ruolo', 'TOT Prev.', 'TOT Eff.'], 1):
        cell = ws.cell(row=tot_row, column=col, value=val)
        cell.font = header_font
        cell.fill = PatternFill('solid', start_color='70AD47')
        cell.alignment = header_align
        cell.border = thin_border
    tot_row += 1

    for role in STANDARD_ROLES:
        ws.cell(row=tot_row, column=1, value=role).font = role_font
        ws.cell(row=tot_row, column=1).border = thin_border
        ws.cell(row=tot_row, column=2, value=role_totals[role]['expected']).border = thin_border
        ws.cell(row=tot_row, column=3, value=role_totals[role]['worked']).border = thin_border
        tot_row += 1

    grand_fill = PatternFill('solid', start_color='C6EFCE')
    ws.cell(row=tot_row, column=1, value="TOTALE").font = Font(bold=True)
    ws.cell(row=tot_row, column=1).fill = grand_fill
    ws.cell(row=tot_row, column=1).border = thin_border
    cell_gt_exp = ws.cell(row=tot_row, column=2, value=grand_total_exp)
    cell_gt_wrk = ws.cell(row=tot_row, column=3, value=grand_total_wrk)
    for c in [cell_gt_exp, cell_gt_wrk]:
        c.font = Font(bold=True)
        c.fill = grand_fill
        c.border = thin_border

    current_row += 3

    current_row = 22

    ws.cell(row=current_row, column=1, value="RIEPILOGO SPRINT").font = title_font
    current_row += 2

    headers2 = ['Sprint', 'Completate', 'In Ritardo', 'Ore Stimate', 'Ore Lavorate', 'Ore Stim. Compl.', 'Ore Lav. Compl.', 'TCR (%)']
    for col, val in enumerate(headers2, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row += 1

    for sprint in sorted_sprints + (["Backlog"] if "Backlog" in sprint_summary else []):
        if sprint in sprint_summary:
            data = sprint_summary[sprint]

            comp = data['completate']
            rit = data['ritardo']
            denominatore = comp + rit

            if denominatore > 0:
                tcr_val = (comp / denominatore) * 100
            else:
                tcr_val = 0

            row_data = [
                sprint,
                data['completate'],
                data['ritardo'],
                data['ore_stimate'],
                data['ore_lavorate'],
                data['ore_stimate_completate'],
                data['ore_lavorate_completate'],
                f"{tcr_val:.1f}%"
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = thin_border
                if col == 1:
                    cell.font = role_font
            current_row += 1

    ws.column_dimensions['A'].width = 25
    for i in range(2, len(header_row) + 1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = 12

    current_row += 2

    ws.cell(row=current_row, column=1, value="ORE PER PERSONA ASSEGNATA").font = title_font
    current_row += 2

    headers3 = ['Persona', 'N. Issue', 'Ore Prev.', 'Ore Eff.', 'Differenza']
    for col, val in enumerate(headers3, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = PatternFill('solid', start_color='ED7D31')
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1

    sorted_assignees = sorted(assignee_hours.items(), key=lambda x: x[1]['worked'], reverse=True)

    assignee_to_anonymous = {}
    person_counter = 1
    for assignee, _ in sorted_assignees:
        if assignee == 'Non assegnato':
            assignee_to_anonymous[assignee] = 'Non assegnato'
        else:
            assignee_to_anonymous[assignee] = f'Persona {person_counter}'
            person_counter += 1

    total_assignee_exp = 0
    total_assignee_wrk = 0
    total_issue_count = 0

    for assignee, hours in sorted_assignees:
        anonymous_name = assignee_to_anonymous[assignee]
        ws.cell(row=current_row, column=1, value=anonymous_name).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=2, value=hours['issue_count']).border = thin_border
        ws.cell(row=current_row, column=3, value=round(hours['expected'], 1)).border = thin_border
        ws.cell(row=current_row, column=4, value=round(hours['worked'], 1)).border = thin_border
        diff = round(hours['worked'] - hours['expected'], 1)
        diff_cell = ws.cell(row=current_row, column=5, value=diff)
        diff_cell.border = thin_border
        if diff < 0:
            diff_cell.font = Font(color='FF0000')
        elif diff > 0:
            diff_cell.font = Font(color='008000')

        total_assignee_exp += hours['expected']
        total_assignee_wrk += hours['worked']
        total_issue_count += hours['issue_count']
        current_row += 1

    ws.cell(row=current_row, column=1, value="TOTALE").font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=2, value=total_issue_count).font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=3, value=round(total_assignee_exp, 1)).font = Font(bold=True)
    ws.cell(row=current_row, column=3).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=3).border = thin_border
    ws.cell(row=current_row, column=4, value=round(total_assignee_wrk, 1)).font = Font(bold=True)
    ws.cell(row=current_row, column=4).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=4).border = thin_border
    ws.cell(row=current_row, column=5, value=round(total_assignee_wrk - total_assignee_exp, 1)).font = Font(bold=True)
    ws.cell(row=current_row, column=5).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=5).border = thin_border

    current_row += 3

    ws.cell(row=current_row, column=1, value="EARNED VALUE MANAGEMENT (€)").font = title_font
    current_row += 2

    def calc_cost(hours_dict):
        cost_exp = 0
        cost_wrk = 0
        for role in STANDARD_ROLES:
            rate = HOURLY_RATES.get(role, 15)
            cost_exp += hours_dict[role]['expected'] * rate
            cost_wrk += hours_dict[role]['worked'] * rate
        return cost_exp, cost_wrk

    evm_data = {
        'BAC': [], 'AC': [], 'PV': [], 'EV': [],
        'PV_acc': [], 'AC_acc': [], 'EV_acc': [],
        'SV': [], 'CV': [], 'EAC': [], 'ETC': []
    }

    pv_accumulated = 0
    ac_accumulated = 0
    ev_accumulated = 0

    for sprint in sorted_sprints:
        pv_cost, ac_cost = calc_cost(sprint_role_hours[sprint])

        sprint_data = sprint_summary.get(sprint, {})
        ore_stimate_totali = sprint_data.get('ore_stimate', 0)
        ore_stimate_completate = sprint_data.get('ore_lavorate', 0)

        ev_cost = 0
        if ore_stimate_totali > 0:
            completion_ratio = ore_stimate_completate / ore_stimate_totali
            ev_cost = pv_cost * completion_ratio

        pv_accumulated += pv_cost
        ac_accumulated += ac_cost
        ev_accumulated += ev_cost

        sv = ev_cost - pv_cost
        cv = ev_cost - ac_cost

        cpi = ev_accumulated / ac_accumulated if ac_accumulated > 0 else 0
        eac = BAT_BUDGET / cpi if cpi > 0 else 0
        etc = eac - ac_accumulated if eac > 0 else 0

        evm_data['BAC'].append(BAT_BUDGET)
        evm_data['AC'].append(round(ac_cost, 2))
        evm_data['PV'].append(round(pv_cost, 2))
        evm_data['EV'].append(round(ev_cost, 2))
        evm_data['PV_acc'].append(round(pv_accumulated, 2))
        evm_data['AC_acc'].append(round(ac_accumulated, 2))
        evm_data['EV_acc'].append(round(ev_accumulated, 2))
        evm_data['SV'].append(round(sv, 2))
        evm_data['CV'].append(round(cv, 2))
        evm_data['EAC'].append(round(eac, 2))
        evm_data['ETC'].append(round(etc, 2))

    evm_header = [''] + [s.replace('Sprint ', 's') for s in sorted_sprints]
    for col, val in enumerate(evm_header, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = PatternFill('solid', start_color='7030A0')
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1

    evm_rows = [
        ('BAC - Budget at Completion (€)', evm_data['BAC']),
        ('AC - Actual Cost (€)', evm_data['AC']),
        ('PV - Planned Value (€)', evm_data['PV']),
        ('EV - Earned Value (€)', evm_data['EV']),
        ('PV acc. (€)', evm_data['PV_acc']),
        ('AC acc. (€)', evm_data['AC_acc']),
        ('EV acc. (€)', evm_data['EV_acc']),
        ('SV - Schedule Variance (€)', evm_data['SV']),
        ('CV - Cost Variance (€)', evm_data['CV']),
        ('EAC - Estimate At Completion (€)', evm_data['EAC']),
        ('ETC - Estimate To Complete (€)', evm_data['ETC']),
    ]

    for row_name, values in evm_rows:
        ws.cell(row=current_row, column=1, value=row_name).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border

        for col, val in enumerate(values, 2):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.border = thin_border
            if 'SV' in row_name or 'CV' in row_name:
                if val < 0:
                    cell.font = Font(color='FF0000')
                elif val > 0:
                    cell.font = Font(color='008000')
        current_row += 1

    
    
    
    current_row += 3

    ws.cell(row=current_row, column=1, value="MPC-PRCT - PULL REQUEST CYCLE TIME").font = title_font
    current_row += 2

    
    from datetime import datetime

    def assign_pr_to_sprint(pr, sprint_dates, sorted_sprints):
        """Assegna una PR a uno sprint in base alla data di merge."""
        merge_str = pr.get('merged_at')
        if not merge_str:
            return None
        try:
            merge_dt = datetime.fromisoformat(merge_str.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            return None

        for sprint_name in sorted_sprints:
            if sprint_name in sprint_dates:
                sd = sprint_dates[sprint_name]
                if sd['start'] <= merge_dt <= sd['end']:
                    return sprint_name
        return None

    sprint_pr_stats = defaultdict(lambda: {'count': 0, 'merged': 0, 'prct_hours': []})

    for pr in pr_data:
        sprint_assigned = assign_pr_to_sprint(pr, sprint_dates, sorted_sprints)
        if not sprint_assigned:
            continue

        sprint_pr_stats[sprint_assigned]['count'] += 1

        if pr.get('merged_at') and pr.get('created_at'):
            sprint_pr_stats[sprint_assigned]['merged'] += 1
            duration_h = calc_duration_days(pr['created_at'], pr['merged_at'])
            if duration_h is not None:
                
                sprint_pr_stats[sprint_assigned]['prct_hours'].append(round(duration_h * 24, 1))

    pr_headers = ['Sprint', 'N. PR', 'PR Merged', 'PRCT Medio (ore)']
    pr_fill = PatternFill('solid', start_color='5B9BD5')
    for col, val in enumerate(pr_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = pr_fill
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1

    total_pr_count = 0
    total_merged = 0
    all_prct = []

    for sprint in sorted_sprints:
        stats = sprint_pr_stats[sprint]
        count = stats['count']
        merged = stats['merged']
        prct_list = stats['prct_hours']
        avg_prct = round(sum(prct_list) / len(prct_list), 1) if prct_list else ''

        total_pr_count += count
        total_merged += merged
        all_prct.extend(prct_list)

        row_vals = [sprint, count, merged, avg_prct]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.border = thin_border
            if col == 1:
                cell.font = role_font
        current_row += 1

    
    total_avg_prct = round(sum(all_prct) / len(all_prct), 1) if all_prct else ''
    summary_fill = PatternFill('solid', start_color='D6E4F0')
    total_vals = ['TOTALE', total_pr_count, total_merged, total_avg_prct]
    for col, val in enumerate(total_vals, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = summary_fill
        cell.border = thin_border

    current_row += 3

    
    
    
    import statistics
    sprint_workload_stddev = {}
    for sprint in sorted_sprints:
        assignees_in_sprint = sprint_assignee_hours.get(sprint, {})
        hours_list = [
            data['worked'] for name, data in assignees_in_sprint.items()
            if name != 'Non assegnato' and data['worked'] > 0
        ]
        if len(hours_list) >= 2:
            total_hours = sum(hours_list)
            if total_hours > 0:
                pct_list = [h / total_hours for h in hours_list]
                stddev = statistics.stdev(pct_list)
                sprint_workload_stddev[sprint] = round(stddev * 100, 2)
            else:
                sprint_workload_stddev[sprint] = None
        else:
            sprint_workload_stddev[sprint] = None

    
    
    
    current_row += 3

    ws.cell(row=current_row, column=1, value="MPC-WSD - DISTRIBUZIONE CARICO ORE").font = title_font
    current_row += 2

    
    wsd_header = ['Persona'] + [s.replace('Sprint ', 'S') for s in sorted_sprints]
    wsd_fill = PatternFill('solid', start_color='2E75B6')
    for col, val in enumerate(wsd_header, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = wsd_fill
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1

    
    all_people = set()
    for sprint in sorted_sprints:
        for name in sprint_assignee_hours.get(sprint, {}):
            if name != 'Non assegnato':
                all_people.add(name)
    sorted_people = sorted(all_people)

    
    for person in sorted_people:
        anon_name = assignee_to_anonymous.get(person, person)
        ws.cell(row=current_row, column=1, value=anon_name).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border

        for i, sprint in enumerate(sorted_sprints):
            col_idx = 2 + i
            assignees_in_sprint = sprint_assignee_hours.get(sprint, {})
            total_sprint_hours = sum(
                d['worked'] for n, d in assignees_in_sprint.items() if n != 'Non assegnato'
            )
            person_hours = assignees_in_sprint.get(person, {}).get('worked', 0)
            if total_sprint_hours > 0 and person_hours > 0:
                pct = round((person_hours / total_sprint_hours) * 100, 1)
                cell = ws.cell(row=current_row, column=col_idx, value=f"{pct}%")
            else:
                cell = ws.cell(row=current_row, column=col_idx, value='-')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

    
    stddev_fill = PatternFill('solid', start_color='D6E4F0')
    ws.cell(row=current_row, column=1, value="STD DEV").font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = stddev_fill
    ws.cell(row=current_row, column=1).border = thin_border

    for i, sprint in enumerate(sorted_sprints):
        col_idx = 2 + i
        val = sprint_workload_stddev.get(sprint)
        if val is not None:
            cell = ws.cell(row=current_row, column=col_idx, value=f"{val}%")
            if val <= 6.30:
                cell.font = Font(bold=True, color='006100')
                cell.fill = PatternFill('solid', start_color='C6EFCE')
            elif val <= 10.0:
                cell.font = Font(bold=True, color='9C6500')
                cell.fill = PatternFill('solid', start_color='FFEB9C')
            else:
                cell.font = Font(bold=True, color='9C0006')
                cell.fill = PatternFill('solid', start_color='FFC7CE')
        else:
            cell = ws.cell(row=current_row, column=col_idx, value='N/A')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    current_row += 1

    
    ws.cell(row=current_row, column=1, value="Soglia Ottimo / Acc.").font = Font(italic=True)
    ws.cell(row=current_row, column=1).border = thin_border
    for i in range(len(sorted_sprints)):
        cell = ws.cell(row=current_row, column=2 + i, value="6.30% / 10%")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(italic=True, color='808080')

    current_row += 3

    
    
    

    ws.cell(row=current_row, column=1, value="MPC-TE - TIME EFFICIENCY").font = title_font
    current_row += 2

    te_header = ['Sprint', 'Ore Effettive', 'Ore Palestra', 'Ore Totali', 'TE (%)', 'Giudizio']
    te_fill = PatternFill('solid', start_color='BF8F00')
    for col, val in enumerate(te_header, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = te_fill
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1

    for sprint in sorted_sprints:
        ore_eff = sprint_summary.get(sprint, {}).get('ore_lavorate', 0)
        ore_pal = sprint_palestra_hours.get(sprint, 0)
        ore_tot = ore_eff + ore_pal
        te_val = round((ore_eff / ore_tot) * 100, 1) if ore_tot > 0 else 0

        if te_val >= 90:
            giudizio = 'Ottimo'
            giud_font = Font(bold=True, color='006100')
        elif te_val >= 50:
            giudizio = 'Accettabile'
            giud_font = Font(bold=True, color='9C6500')
        else:
            giudizio = 'Non acc.'
            giud_font = Font(bold=True, color='9C0006')

        ws.cell(row=current_row, column=1, value=sprint).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=2, value=ore_eff).border = thin_border
        ws.cell(row=current_row, column=3, value=ore_pal).border = thin_border
        ws.cell(row=current_row, column=4, value=ore_tot).border = thin_border
        te_cell = ws.cell(row=current_row, column=5, value=f"{te_val}%")
        te_cell.border = thin_border
        te_cell.alignment = Alignment(horizontal='center')
        giud_cell = ws.cell(row=current_row, column=6, value=giudizio)
        giud_cell.border = thin_border
        giud_cell.font = giud_font
        current_row += 1

    current_row += 3

    
    
    

    ws.cell(row=current_row, column=1, value="MPC-QMS - QUALITY METRICS SATISFIED").font = title_font
    current_row += 2

    
    
    qms_header = ['ID', 'Nome Metrica', 'Accettabile', 'Ottimo'] + \
                 [s.replace('Sprint ', 'S') for s in sorted_sprints] + ['QMS (%)']
    qms_fill = PatternFill('solid', start_color='548235')
    for col, val in enumerate(qms_header, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = qms_fill
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1

    ok_fill = PatternFill('solid', start_color='C6EFCE')
    warn_fill = PatternFill('solid', start_color='FFEB9C')
    fail_fill = PatternFill('solid', start_color='FFC7CE')
    optimal_font = Font(color='006100')
    acceptable_font = Font(color='9C6500')
    fail_font = Font(color='9C0006')

    
    sprint_tcr = {}
    sprint_ts = {}
    for sprint in sorted_sprints:
        data = sprint_summary.get(sprint, {})
        comp = data.get('completate', 0)
        rit = data.get('ritardo', 0)
        denom = comp + rit
        sprint_tcr[sprint] = (comp / denom) * 100 if denom > 0 else 0
        
        sprint_ts[sprint] = (rit / denom) * 100 if denom > 0 else 0

    
    sprint_prct_avg = {}
    for sprint in sorted_sprints:
        stats = sprint_pr_stats.get(sprint, {'prct_hours': []})
        prct_list = stats['prct_hours']
        sprint_prct_avg[sprint] = round(sum(prct_list) / len(prct_list), 1) if prct_list else None

    
    
    def make_qms_metrics():
        metrics = []

        
        metrics.append({
            'id': 'MPC-PV', 'name': 'Planned Value',
            'acc_str': '> 0€', 'opt_str': 'Pianificato',
            'values': evm_data['PV'],
            'check_acc': lambda v: v > 0,
            'check_opt': lambda v: v > 0,
        })
        
        metrics.append({
            'id': 'MPC-AC', 'name': 'Actual Cost',
            'acc_str': '> 0€', 'opt_str': '≤ EAC',
            'values': evm_data['AC'],
            'check_acc': lambda v: v > 0,
            'check_opt': lambda v, i=None: True,  
        })
        
        metrics.append({
            'id': 'MPC-EV', 'name': 'Earned Value',
            'acc_str': '> 0€', 'opt_str': '> PV',
            'values': evm_data['EV'],
            'check_acc': lambda v: v > 0,
            'check_opt': lambda v, i=None: True,
        })
        
        metrics.append({
            'id': 'MPC-BAC', 'name': 'Budget At Completion',
            'acc_str': f'{BAT_BUDGET}€', 'opt_str': f'{BAT_BUDGET}€',
            'values': evm_data['BAC'],
            'check_acc': lambda v: v == BAT_BUDGET,
            'check_opt': lambda v: v == BAT_BUDGET,
        })
        
        metrics.append({
            'id': 'MPC-EAC', 'name': 'Estimated At Completion',
            'acc_str': f'≥ {BAT_BUDGET * 0.95:.0f}€', 'opt_str': f'{BAT_BUDGET}€',
            'values': evm_data['EAC'],
            'check_acc': lambda v: v >= BAT_BUDGET * 0.95,
            'check_opt': lambda v: abs(v - BAT_BUDGET) < BAT_BUDGET * 0.02,
        })
        
        metrics.append({
            'id': 'MPC-ETC', 'name': 'Estimated To Complete',
            'acc_str': '≥ 0€', 'opt_str': '≥ 0€',
            'values': evm_data['ETC'],
            'check_acc': lambda v: v >= 0,
            'check_opt': lambda v: v >= 0,
        })
        
        metrics.append({
            'id': 'MPC-CV', 'name': 'Cost Variance',
            'acc_str': '> 0€', 'opt_str': '0€',
            'values': evm_data['CV'],
            'check_acc': lambda v: v > 0,
            'check_opt': lambda v: v >= 0,
        })
        
        metrics.append({
            'id': 'MPC-SV', 'name': 'Schedule Variance',
            'acc_str': '≥ 0€', 'opt_str': '> 0€',
            'values': evm_data['SV'],
            'check_acc': lambda v: v >= 0,
            'check_opt': lambda v: v > 0,
        })
        
        tcr_vals = [round(sprint_tcr.get(s, 0), 1) for s in sorted_sprints]
        metrics.append({
            'id': 'MPC-TCR', 'name': 'Task Completion Rate',
            'acc_str': '≥ 85%', 'opt_str': '100%',
            'values': tcr_vals,
            'check_acc': lambda v: v >= 85,
            'check_opt': lambda v: v >= 100,
            'is_pct': True,
        })
        
        ts_vals = [round(sprint_ts.get(s, 0), 1) for s in sorted_sprints]
        metrics.append({
            'id': 'MPC-TS', 'name': 'Task Slippage',
            'acc_str': '≤ 15%', 'opt_str': '0%',
            'values': ts_vals,
            'check_acc': lambda v: v <= 15,
            'check_opt': lambda v: v == 0,
            'is_pct': True,
        })
        
        prct_vals = [sprint_prct_avg.get(s) for s in sorted_sprints]
        metrics.append({
            'id': 'MTC-PRCT', 'name': 'PR Cycle Time',
            'acc_str': '≤ 48 ore', 'opt_str': '≤ 24 ore',
            'values': prct_vals,
            'check_acc': lambda v: v is not None and v <= 48,
            'check_opt': lambda v: v is not None and v <= 24,
        })

        
        wsd_vals = [sprint_workload_stddev.get(s) for s in sorted_sprints]
        metrics.append({
            'id': 'MPC-WSD', 'name': 'Workload Std Dev',
            'acc_str': '≤ 10%', 'opt_str': '≤ 6.30%',
            'values': wsd_vals,
            'check_acc': lambda v: v is not None and v <= 10.0,
            'check_opt': lambda v: v is not None and v <= 6.30,
            'is_pct': True,
        })

        
        
        te_vals = []
        for s in sorted_sprints:
            ore_eff = sprint_summary.get(s, {}).get('ore_lavorate', 0)
            ore_pal = sprint_palestra_hours.get(s, 0)
            ore_totali = ore_eff + ore_pal
            if ore_totali > 0:
                te_vals.append(round((ore_eff / ore_totali) * 100, 1))
            else:
                te_vals.append(None)
        metrics.append({
            'id': 'MPC-TE', 'name': 'Time Efficiency',
            'acc_str': '≥ 80%', 'opt_str': '≥ 90%',
            'values': te_vals,
            'check_acc': lambda v: v is not None and v >= 80,
            'check_opt': lambda v: v is not None and v >= 90,
            'is_pct': True,
        })

        return metrics

    qms_metrics = make_qms_metrics()

    
    sprint_acc_count = [0] * len(sorted_sprints)
    sprint_total_count = [0] * len(sorted_sprints)

    for m in qms_metrics:
        ws.cell(row=current_row, column=1, value=m['id']).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=2, value=m['name']).border = thin_border
        ws.cell(row=current_row, column=3, value=m['acc_str']).border = thin_border
        ws.cell(row=current_row, column=4, value=m['opt_str']).border = thin_border

        metric_acc_count = 0
        metric_total = 0

        for i, val in enumerate(m['values']):
            col_idx = 5 + i
            display_val = val
            if val is None:
                display_val = 'N/A'
            elif m.get('is_pct'):
                display_val = f"{val:.1f}%"

            cell = ws.cell(row=current_row, column=col_idx, value=display_val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            if val is not None:
                metric_total += 1
                sprint_total_count[i] += 1
                is_opt = m['check_opt'](val)
                is_acc = m['check_acc'](val)

                if is_opt:
                    cell.fill = ok_fill
                    cell.font = optimal_font
                    metric_acc_count += 1
                    sprint_acc_count[i] += 1
                elif is_acc:
                    cell.fill = warn_fill
                    cell.font = acceptable_font
                    metric_acc_count += 1
                    sprint_acc_count[i] += 1
                else:
                    cell.fill = fail_fill
                    cell.font = fail_font

        
        qms_pct = round((metric_acc_count / metric_total) * 100, 1) if metric_total > 0 else 0
        qms_cell = ws.cell(row=current_row, column=5 + len(sorted_sprints), value=f"{qms_pct:.1f}%")
        qms_cell.border = thin_border
        qms_cell.alignment = Alignment(horizontal='center')
        qms_cell.font = Font(bold=True)
        if qms_pct >= 85:
            qms_cell.fill = ok_fill
        elif qms_pct >= 50:
            qms_cell.fill = warn_fill
        else:
            qms_cell.fill = fail_fill

        current_row += 1

    
    qms_total_fill = PatternFill('solid', start_color='D9E2F3')
    ws.cell(row=current_row, column=1, value="QMS").font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=1).fill = qms_total_fill
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=2, value="Quality Metrics Satisfied").font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = qms_total_fill
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=3).fill = qms_total_fill
    ws.cell(row=current_row, column=3).border = thin_border
    ws.cell(row=current_row, column=4).fill = qms_total_fill
    ws.cell(row=current_row, column=4).border = thin_border

    all_acc = 0
    all_total = 0
    for i in range(len(sorted_sprints)):
        col_idx = 5 + i
        qms_val = round((sprint_acc_count[i] / sprint_total_count[i]) * 100, 1) if sprint_total_count[i] > 0 else 0
        all_acc += sprint_acc_count[i]
        all_total += sprint_total_count[i]

        cell = ws.cell(row=current_row, column=col_idx, value=f"{qms_val:.1f}%")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        cell.fill = qms_total_fill
        if qms_val >= 85:
            cell.font = Font(bold=True, color='006100')
        elif qms_val >= 50:
            cell.font = Font(bold=True, color='9C6500')
        else:
            cell.font = Font(bold=True, color='9C0006')

    
    global_qms = round((all_acc / all_total) * 100, 1) if all_total > 0 else 0
    gqms_cell = ws.cell(row=current_row, column=5 + len(sorted_sprints), value=f"{global_qms:.1f}%")
    gqms_cell.border = thin_border
    gqms_cell.alignment = Alignment(horizontal='center')
    gqms_cell.font = Font(bold=True, size=12)
    gqms_cell.fill = qms_total_fill

    current_row += 3

    wb.save(OUTPUT_FILE)
    print(f"-> {OUTPUT_FILE}")


if __name__ == "__main__":
    main()