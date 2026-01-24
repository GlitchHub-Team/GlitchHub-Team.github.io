import requests
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time

GITHUB_TOKEN = "" #
ORGANIZATION = "GlitchHub-Team"
PROJECT_NUMBER = 2
OUTPUT_FILE = "report_sprint_gerarchico.xlsx"

EXPECTED_HOURS_FIELD = "Expected Worked Hours"
WORKED_HOURS_FIELD = "Worked hours"
ROLE_FIELD_NAME = "Sprint role"

EXCLUDED_LABELS = ["task-palestra"]

STANDARD_ROLES = [
    "Responsabile",
    "Amministratore",
    "Analista",
    "Progettista",
    "Programmatore",
    "Verificatore"
]

HOURLY_RATES = {
    "Responsabile": 30,
    "Amministratore": 20,
    "Analista": 25,
    "Progettista": 25,
    "Programmatore": 15,
    "Verificatore": 15,
    "Non assegnato": 15
}

BAT_BUDGET = 12975


def get_all_project_items(org, project_num, token):
    url = "https://api.github.com/graphql"
    headers = {"Authorization": f"bearer {token}"}

    all_nodes = []
    has_next_page = True
    end_cursor = None

    while has_next_page:
        after_arg = f', after: "{end_cursor}"' if end_cursor else ""

        query = """
        query {
          organization(login: "%s") {
            projectV2(number: %d) {
              items(first: 100%s) {
                pageInfo { hasNextPage, endCursor }
                nodes {
                  content {
                    ... on Issue {
                      title
                      number
                      state
                      url
                      createdAt
                      author { login }
                      assignees(first: 10) {
                        nodes { login }
                      }
                      labels(first: 10) { nodes { name } }
                    }
                  }
                  fieldValues(first: 20) {
                    nodes {
                      ... on ProjectV2ItemFieldIterationValue {
                        title
                        field { ... on ProjectV2IterationField { name } }
                      }
                      ... on ProjectV2ItemFieldSingleSelectValue {
                        name
                        field { ... on ProjectV2SingleSelectField { name } }
                      }
                      ... on ProjectV2ItemFieldNumberValue {
                        number
                        field { ... on ProjectV2Field { name } }
                      }
                      ... on ProjectV2ItemFieldTextValue {
                        text
                        field { ... on ProjectV2Field { name } }
                      }
                    }
                  }
                }
              }
            }
          }
        }
        """ % (org, project_num, after_arg)

        response = requests.post(url, json={'query': query}, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Errore HTTP: {response.status_code} - {response.text}")

        data = response.json()
        if 'errors' in data:
            raise Exception(data['errors'][0]['message'])

        chunk = data['data']['organization']['projectV2']['items']
        all_nodes.extend(chunk['nodes'])
        has_next_page = chunk['pageInfo']['hasNextPage']
        end_cursor = chunk['pageInfo']['endCursor']
        time.sleep(0.2)

    return all_nodes


def main():
    try:
        items = get_all_project_items(ORGANIZATION, PROJECT_NUMBER, GITHUB_TOKEN)
    except Exception:
        return

    sprint_summary = {}
    role_hours_expected = defaultdict(float)
    role_hours_worked = defaultdict(float)
    sprint_role_hours = defaultdict(lambda: defaultdict(lambda: {'expected': 0, 'worked': 0}))
    author_hours_expected = defaultdict(float)
    author_hours_worked = defaultdict(float)
    sprint_author_hours = defaultdict(lambda: defaultdict(lambda: {'expected': 0, 'worked': 0}))

    assignee_hours = defaultdict(lambda: {'expected': 0, 'worked': 0, 'issue_count': 0})
    sprint_assignee_hours = defaultdict(lambda: defaultdict(lambda: {'expected': 0, 'worked': 0, 'issue_count': 0}))

    for item in items:
        if not item['content']:
            continue

        labels = [l['name'].lower() for l in item['content']['labels']['nodes']]

        if any(excl.lower() in labels for excl in EXCLUDED_LABELS):
            continue

        sprint_name = "Backlog"
        expected_hours = 0
        worked_hours = 0
        role = "Non assegnato"

        for fv in item['fieldValues']['nodes']:
            if not fv or 'field' not in fv:
                continue

            field_name = fv['field'].get('name', '')

            if field_name == 'Sprint':
                sprint_name = fv.get('title') or fv.get('name') or "Backlog"

            if field_name == EXPECTED_HOURS_FIELD:
                expected_hours = fv.get('number') or 0

            if field_name == WORKED_HOURS_FIELD:
                worked_hours = fv.get('number') or 0

            if field_name == ROLE_FIELD_NAME:
                role = fv.get('name') or fv.get('text') or "Non assegnato"

        author = item['content'].get('author', {})
        author_login = author.get('login', 'Sconosciuto') if author else 'Sconosciuto'

        state = item['content']['state']

        if sprint_name not in sprint_summary:
            sprint_summary[sprint_name] = {
                'completate': 0,
                'ritardo': 0,
                'ore_stimate': 0,
                'ore_lavorate': 0,
                'ore_stimate_completate': 0,
                'ore_lavorate_completate': 0
            }

        sprint_summary[sprint_name]['ore_stimate'] += expected_hours
        sprint_summary[sprint_name]['ore_lavorate'] += worked_hours

        if state == 'CLOSED':
            sprint_summary[sprint_name]['completate'] += 1
            sprint_summary[sprint_name]['ore_stimate_completate'] += expected_hours
            sprint_summary[sprint_name]['ore_lavorate_completate'] += worked_hours

        if state == 'CLOSED' and 'delayed' in labels:
            sprint_summary[sprint_name]['ritardo'] += 1

        role_hours_expected[role] += expected_hours
        role_hours_worked[role] += worked_hours
        sprint_role_hours[sprint_name][role]['expected'] += expected_hours
        sprint_role_hours[sprint_name][role]['worked'] += worked_hours

        author_hours_expected[author_login] += expected_hours
        author_hours_worked[author_login] += worked_hours
        sprint_author_hours[sprint_name][author_login]['expected'] += expected_hours
        sprint_author_hours[sprint_name][author_login]['worked'] += worked_hours

        assignees = item['content'].get('assignees', {}).get('nodes', [])
        if assignees:
            num_assignees = len(assignees)
            hours_per_assignee_exp = expected_hours / num_assignees if num_assignees > 0 else 0
            hours_per_assignee_wrk = worked_hours / num_assignees if num_assignees > 0 else 0

            for assignee in assignees:
                assignee_login = assignee.get('login', 'Sconosciuto')
                assignee_hours[assignee_login]['expected'] += hours_per_assignee_exp
                assignee_hours[assignee_login]['worked'] += hours_per_assignee_wrk
                assignee_hours[assignee_login]['issue_count'] += 1
                sprint_assignee_hours[sprint_name][assignee_login]['expected'] += hours_per_assignee_exp
                sprint_assignee_hours[sprint_name][assignee_login]['worked'] += hours_per_assignee_wrk
                sprint_assignee_hours[sprint_name][assignee_login]['issue_count'] += 1
        else:
            assignee_hours['Non assegnato']['expected'] += expected_hours
            assignee_hours['Non assegnato']['worked'] += worked_hours
            assignee_hours['Non assegnato']['issue_count'] += 1
            sprint_assignee_hours[sprint_name]['Non assegnato']['expected'] += expected_hours
            sprint_assignee_hours[sprint_name]['Non assegnato']['worked'] += worked_hours
            sprint_assignee_hours[sprint_name]['Non assegnato']['issue_count'] += 1

    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    role_font = Font(bold=True)

    ws = wb.active
    ws.title = 'Report'

    def sprint_sort_key(name):
        import re
        match = re.search(r'(\d+)', name)
        return int(match.group(1)) if match else 999

    sorted_sprints = sorted(
        [s for s in sprint_role_hours.keys() if s != "Backlog"],
        key=sprint_sort_key
    )

    current_row = 1

    title_font = Font(bold=True, size=14)
    ws.cell(row=current_row, column=1, value="RIEPILOGO ORE PER RUOLO").font = title_font
    current_row += 2

    header_row = ['Ruolo']
    for sprint in sorted_sprints:
        sprint_short = sprint.replace("Sprint ", "S")
        header_row.extend([f'{sprint_short} Prev.', f'{sprint_short} Eff.'])

    for col, val in enumerate(header_row, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row += 1

    role_totals = {}
    for role in STANDARD_ROLES:
        ws.cell(row=current_row, column=1, value=role).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border
        col = 2
        role_total_exp = 0
        role_total_wrk = 0
        for sprint in sorted_sprints:
            exp = sprint_role_hours[sprint][role]['expected']
            wrk = sprint_role_hours[sprint][role]['worked']
            role_total_exp += exp
            role_total_wrk += wrk
            ws.cell(row=current_row, column=col, value=exp).border = thin_border
            ws.cell(row=current_row, column=col+1, value=wrk).border = thin_border
            col += 2
        role_totals[role] = {'expected': role_total_exp, 'worked': role_total_wrk}
        current_row += 1

    total_fill = PatternFill('solid', start_color='D9E2F3')
    ws.cell(row=current_row, column=1, value="TOTALE").font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = total_fill
    ws.cell(row=current_row, column=1).border = thin_border
    col = 2
    grand_total_exp = 0
    grand_total_wrk = 0
    for sprint in sorted_sprints:
        total_exp = sum(sprint_role_hours[sprint][r]['expected'] for r in STANDARD_ROLES)
        total_wrk = sum(sprint_role_hours[sprint][r]['worked'] for r in STANDARD_ROLES)
        grand_total_exp += total_exp
        grand_total_wrk += total_wrk
        cell_exp = ws.cell(row=current_row, column=col, value=total_exp)
        cell_wrk = ws.cell(row=current_row, column=col+1, value=total_wrk)
        for c in [cell_exp, cell_wrk]:
            c.font = Font(bold=True)
            c.fill = total_fill
            c.border = thin_border
        col += 2

    tot_row = 11
    ws.cell(row=tot_row, column=1, value="TOTALI").font = title_font
    tot_row += 1

    for col, val in enumerate(['Ruolo', 'TOT Prev.', 'TOT Eff.'], 1):
        cell = ws.cell(row=tot_row, column=col, value=val)
        cell.font = header_font
        cell.fill = PatternFill('solid', start_color='70AD47')
        cell.alignment = header_align
        cell.border = thin_border
    tot_row += 1

    for role in STANDARD_ROLES:
        ws.cell(row=tot_row, column=1, value=role).font = role_font
        ws.cell(row=tot_row, column=1).border = thin_border
        ws.cell(row=tot_row, column=2, value=role_totals[role]['expected']).border = thin_border
        ws.cell(row=tot_row, column=3, value=role_totals[role]['worked']).border = thin_border
        tot_row += 1

    grand_fill = PatternFill('solid', start_color='C6EFCE')
    ws.cell(row=tot_row, column=1, value="TOTALE").font = Font(bold=True)
    ws.cell(row=tot_row, column=1).fill = grand_fill
    ws.cell(row=tot_row, column=1).border = thin_border
    cell_gt_exp = ws.cell(row=tot_row, column=2, value=grand_total_exp)
    cell_gt_wrk = ws.cell(row=tot_row, column=3, value=grand_total_wrk)
    for c in [cell_gt_exp, cell_gt_wrk]:
        c.font = Font(bold=True)
        c.fill = grand_fill
        c.border = thin_border

    current_row += 3

    current_row = 22

    ws.cell(row=current_row, column=1, value="RIEPILOGO SPRINT").font = title_font
    current_row += 2

    headers2 = ['Sprint', 'Completate', 'In Ritardo', 'Ore Stimate', 'Ore Lavorate', 'Ore Stim. Compl.', 'Ore Lav. Compl.', 'TCR (%)']
    for col, val in enumerate(headers2, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row += 1

    for sprint in sorted_sprints + (["Backlog"] if "Backlog" in sprint_summary else []):
        if sprint in sprint_summary:

            data = sprint_summary[sprint]

            comp = data['completate']
            rit = data['ritardo']
            denominatore = comp + rit
            
            if denominatore > 0:
                tcr_val = (comp / denominatore) * 100
            else:
                tcr_val = 0

            row_data = [
                sprint,
                data['completate'],
                data['ritardo'],
                data['ore_stimate'],
                data['ore_lavorate'],
                data['ore_stimate_completate'],
                data['ore_lavorate_completate'],
                f"{tcr_val:.1f}%"
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = thin_border
                if col == 1:
                    cell.font = role_font
            current_row += 1

    ws.column_dimensions['A'].width = 25
    for i in range(2, len(header_row) + 1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = 12

    current_row += 2

    ws.cell(row=current_row, column=1, value="ORE PER PERSONA ASSEGNATA").font = title_font
    current_row += 2

    headers3 = ['Persona', 'N. Issue', 'Ore Prev.', 'Ore Eff.', 'Differenza']
    for col, val in enumerate(headers3, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = PatternFill('solid', start_color='ED7D31')
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1

    sorted_assignees = sorted(assignee_hours.items(), key=lambda x: x[1]['worked'], reverse=True)

    assignee_to_anonymous = {}
    person_counter = 1
    for assignee, _ in sorted_assignees:
        if assignee == 'Non assegnato':
            assignee_to_anonymous[assignee] = 'Non assegnato'
        else:
            assignee_to_anonymous[assignee] = f'Persona {person_counter}'
            person_counter += 1

    total_assignee_exp = 0
    total_assignee_wrk = 0
    total_issue_count = 0

    for assignee, hours in sorted_assignees:
        anonymous_name = assignee_to_anonymous[assignee]
        ws.cell(row=current_row, column=1, value=anonymous_name).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=2, value=hours['issue_count']).border = thin_border
        ws.cell(row=current_row, column=3, value=round(hours['expected'], 1)).border = thin_border
        ws.cell(row=current_row, column=4, value=round(hours['worked'], 1)).border = thin_border
        diff = round(hours['worked'] - hours['expected'], 1)
        diff_cell = ws.cell(row=current_row, column=5, value=diff)
        diff_cell.border = thin_border
        if diff < 0:
            diff_cell.font = Font(color='FF0000')
        elif diff > 0:
            diff_cell.font = Font(color='008000')

        total_assignee_exp += hours['expected']
        total_assignee_wrk += hours['worked']
        total_issue_count += hours['issue_count']
        current_row += 1

    ws.cell(row=current_row, column=1, value="TOTALE").font = Font(bold=True)
    ws.cell(row=current_row, column=1).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=2, value=total_issue_count).font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=3, value=round(total_assignee_exp, 1)).font = Font(bold=True)
    ws.cell(row=current_row, column=3).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=3).border = thin_border
    ws.cell(row=current_row, column=4, value=round(total_assignee_wrk, 1)).font = Font(bold=True)
    ws.cell(row=current_row, column=4).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=4).border = thin_border
    ws.cell(row=current_row, column=5, value=round(total_assignee_wrk - total_assignee_exp, 1)).font = Font(bold=True)
    ws.cell(row=current_row, column=5).fill = PatternFill('solid', start_color='FCE4D6')
    ws.cell(row=current_row, column=5).border = thin_border

    current_row += 3
    
    ws.cell(row=current_row, column=1, value="EARNED VALUE MANAGEMENT (€)").font = title_font
    current_row += 2

    def calc_cost(hours_dict):
        cost_exp = 0
        cost_wrk = 0
        for role in STANDARD_ROLES:
            rate = HOURLY_RATES.get(role, 15)
            cost_exp += hours_dict[role]['expected'] * rate
            cost_wrk += hours_dict[role]['worked'] * rate
        return cost_exp, cost_wrk

    evm_data = {
        'BAC': [],
        'AC': [],
        'PV': [],
        'EV': [],
        'PV_acc': [],
        'AC_acc': [],
        'EV_acc': [],
        'SV': [],
        'CV': [],
        'EAC': [],
        'ETC': []
    }

    pv_accumulated = 0
    ac_accumulated = 0
    ev_accumulated = 0

    for sprint in sorted_sprints:
        pv_cost, ac_cost = calc_cost(sprint_role_hours[sprint])

        sprint_data = sprint_summary.get(sprint, {})
        ore_stimate_totali = sprint_data.get('ore_stimate', 0)
        ore_stimate_completate = sprint_data.get('ore_stimate_completate', 0)

        ev_cost = 0
        if ore_stimate_totali > 0:
            completion_ratio = ore_stimate_completate / ore_stimate_totali
            ev_cost = pv_cost * completion_ratio

        pv_accumulated += pv_cost
        ac_accumulated += ac_cost
        ev_accumulated += ev_cost

        sv = ev_cost - pv_cost
        cv = ev_cost - ac_cost

        # CPI = EV_acc / AC_acc
        cpi = ev_accumulated / ac_accumulated if ac_accumulated > 0 else 0

        # EAC = BAC / CPI
        eac = BAT_BUDGET / cpi if cpi > 0 else 0

        # ETC = EAC - AC_acc 
        etc = eac - ac_accumulated if eac > 0 else 0

        evm_data['BAC'].append(BAT_BUDGET)
        evm_data['AC'].append(round(ac_cost, 2))
        evm_data['PV'].append(round(pv_cost, 2))
        evm_data['EV'].append(round(ev_cost, 2))
        evm_data['PV_acc'].append(round(pv_accumulated, 2))
        evm_data['AC_acc'].append(round(ac_accumulated, 2))
        evm_data['EV_acc'].append(round(ev_accumulated, 2))
        evm_data['SV'].append(round(sv, 2))
        evm_data['CV'].append(round(cv, 2))
        evm_data['EAC'].append(round(eac, 2))
        evm_data['ETC'].append(round(etc, 2))

    evm_header = [''] + [s.replace('Sprint ', 's') for s in sorted_sprints]
    for col, val in enumerate(evm_header, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = header_font
        cell.fill = PatternFill('solid', start_color='7030A0')
        cell.alignment = header_align
        cell.border = thin_border
    current_row += 1

    evm_rows = [
        ('BAC - Budget at Completion (€)', evm_data['BAC']),
        ('AC - Actual Cost (€)', evm_data['AC']),
        ('PV - Planned Value (€)', evm_data['PV']),
        ('EV - Earned Value (€)', evm_data['EV']),
        ('PV acc. (€)', evm_data['PV_acc']),
        ('AC acc. (€)', evm_data['AC_acc']),
        ('EV acc. (€)', evm_data['EV_acc']),
        ('SV - Schedule Variance (€)', evm_data['SV']),
        ('CV - Cost Variance (€)', evm_data['CV']),
        ('EAC - Estimate At Completion (€)', evm_data['EAC']),
        ('ETC - Estimate To Complete (€)', evm_data['ETC']),
    ]

    for row_name, values in evm_rows:
        ws.cell(row=current_row, column=1, value=row_name).font = role_font
        ws.cell(row=current_row, column=1).border = thin_border

        for col, val in enumerate(values, 2):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.border = thin_border
            if 'SV' in row_name or 'CV' in row_name:
                if val < 0:
                    cell.font = Font(color='FF0000')
                elif val > 0:
                    cell.font = Font(color='008000')
        current_row += 1

    current_row += 3

    wb.save(OUTPUT_FILE)
    print(f"-> {OUTPUT_FILE}")


if __name__ == "__main__":
    main()